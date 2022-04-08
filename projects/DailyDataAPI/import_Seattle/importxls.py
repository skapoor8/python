# importxls.py
# written by: Thomas A. Grate
# copyright (c) 2017 by Thomas A. Grate, all rights reserved.
#
# Refreshes the OYD Daily Program Database for Testing

import sqlite3
from db import *
from school import *
from region import *
from natarea import *
from student import *
from print_table import *
from openpyxl import load_workbook
from datetime import date

class Students_Table_Management (object):
    """ Student_Table object
        Management functions for 'students' Table """

    def __init__(self):
        self.student = Student()

    def drop(self, db):
        db.cursor.execute('DROP TABLE IF EXISTS students;')

    def create(self, db):
        db.cursor.execute('CREATE TABLE students (' + self.student.schema_create + ')')

def open_xls(filename):
    # Load in the workbook
    wb = load_workbook(filename)

    # Get sheet names
    print(wb.get_sheet_names())

    # Get a sheet by name
    sheet = wb.get_sheet_by_name('Students')

    # Print the sheet title
    print(sheet.title)

    # Print row per row
    for cellObj in sheet['A5':'BH5']:
      for cell in cellObj:
              print(cell.coordinate, cell.value)
      print('--- END ---')

    print(f"Start Row = 5, End Row = {sheet.max_row}")

    return sheet

# convert an Excel date in ordinal form to a Human Readable Date
def convert_xls_date (xl_date_ord):
    if xl_date_ord is not None:
        # adjust for excel vs python ordinal represenatation
        py_date_ord = int(xl_date_ord) + 693594

        # convert ordinal to a pyton date object
        py_date = date.fromordinal(py_date_ord)

        # construct a human readable date to store in db
        txt_date = str(py_date.month) + '/' + str(py_date.day) + '/' + \
            str(py_date.year)

        return txt_date
    else:
        return ''

def import_students (dbname, sheet):
    # Open the database designated by dbname
    database = Database()
    database.open_db(dbname)

    # Instantiate a Table object
    stm = Students_Table_Management()

    # Delete the table if it already exists
    stm.drop (database)

    # Create students table
    stm.create (database)

    # Add each student in filename to the student table in dbname
    print(f"Importing {sheet.max_row - 4} rows")

    for i in range (5, sheet.max_row + 1):
        student = Student()

        oyd_id = sheet.cell(row=i, column=2).value
        if oyd_id is None: oyd_id = ''
        student.attrs['oyd_id'] = oyd_id

        student.attrs['first_name'] = sheet.cell(row=i, column=3).value

        middle_name = sheet.cell(row=i, column=4).value
        if middle_name == None: middle_name = ''
        student.attrs['middle_name'] = middle_name

        # check last name is not None, otherwise, give them NoLastName
        last_name = sheet.cell(row=i, column=5).value
        if last_name is None:
            last_name = "NoLastName"
        student.attrs['last_name'] = last_name

        nickname = sheet.cell(row=i, column=6).value
        if nickname is None: nickname = ''
        student.attrs['nickname'] = nickname

        # check that the age is not 117, change to None
        age = sheet.cell(row=i, column=17).value
        if age == 117:
            student.attrs['age'] = 0
        elif age is None:
            student.attrs['age'] = 0
        else:
            student.attrs['age'] = age

        student.attrs['birth_date'] = convert_xls_date(sheet.cell(row=i, column=16).value)

        student.attrs['school'] = sheet.cell(row=i, column=11).value

        student.attrs['start_date'] = convert_xls_date(sheet.cell(row=i, column=7).value)

        status = sheet.cell(row=i, column=8).value
        status = status.lower()
        student.attrs['status'] = status

#        student.attrs['drop_date'] = convert_xls_date(sheet.cell(row=i, column=10).value)

        student.attrs['drop_reason'] = sheet.cell(row=i, column=53).value

        position = sheet.cell(row=i, column=12).value
        if position == '-' or position == None:
            position = ''
        student.attrs['position'] = position

        student.attrs['rank'] = sheet.cell(row=i, column=13).value.upper()
        student.attrs['next_rank'] = sheet.cell(row=i, column=14).value.upper()
        student.attrs['class_group'] = sheet.cell(row=i, column=15).value
        student.attrs['street'] = sheet.cell(row=i, column=37).value
        student.attrs['street2'] = ''
        student.attrs['city'] = sheet.cell(row=i, column=38).value
        student.attrs['state'] = sheet.cell(row=i, column=39).value
        student.attrs['postal'] = sheet.cell(row=i, column=40).value
        student.attrs['country'] = "USA"

        email = sheet.cell(row=i, column=44).value
        if email is None: email = ''
        student.attrs['email'] = email

        mobile_phone = sheet.cell(row=i, column=41).value
        if mobile_phone is None: mobile_phone = ''
        student.attrs['mobile_phone'] = mobile_phone

        home_phone = sheet.cell(row=i, column=42).value
        if home_phone is None: home_phone = ''
        student.attrs['home_phone'] = home_phone

        student.attrs['parental_contact'] = sheet.cell(row=i, column=43).value

        last_test_date = sheet.cell(row=i, column=24).value
        if last_test_date == 'unknown':
            last_test_date = ''
        if last_test_date != '':
            student.attrs['last_test_date'] = convert_xls_date(last_test_date)

        next_test_date = sheet.cell(row=i, column=25).value
        if next_test_date == 'unknown':
            next_test_date = ''
        if next_test_date != '':
            student.attrs['next_test_date'] = convert_xls_date(next_test_date)

        print(f"Importing: {sheet.cell(row=i, column=1).value}: {student.attrs['first_name']} {student.attrs['middle_name']} {student.attrs['last_name']}")

        student.put(database)

    # Print the updated students table
    title = f"Imported {filename} into {dbname} Students Table: Added Students to Roster"
    # print_table (database, 'students', title)

    database.close_db()

if __name__ == "__main__":
    # if run as the main program - refresh the roster.db
    dbname = '../oyd_daily.db'
    filename = 'students.xlsx'

    print (f"Importing students from {filename} to {dbname} database file!")
    print ("WARNING: BE VERY CAREFUL IN THE USE OF THIS APP")
    print (f"It WILL delete all STUDENT DATA in the {dbname} database")
    print ("-------------------------------------------------")
    print ("Fix up the XLS:")
    print ("- copy students sheet to new wb 'by value'")
    print ("- convert date formats (3) to no formatting")
    print ('')

    # Students Table
    if  input("Import Student Roster to Databse: Y or n > ") == 'Y':
        print(f"Importing Students from {filename} to Database")

        # get the students sheet from the xls
        sheet = open_xls(filename)

        # import the sheet into the students table
        import_students (dbname, sheet)
    else:
        print (f"OK, leaving the {dbname} database untouched!")
        print('')
